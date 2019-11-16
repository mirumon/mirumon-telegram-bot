from io import BytesIO
from typing import List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.writer.excel import save_virtual_workbook

from app.schemas.mirumon_responses import Software


def init_file_columns(worksheet: Worksheet) -> None:
    worksheet.cell(row=1, column=1).value = "NAME"  # noqa WPS 110
    worksheet.cell(row=1, column=2).value = "VENDOR"  # noqa WPS 110
    worksheet.cell(row=1, column=3).value = "VERSION"  # noqa WPS 110


def get_excel_file(programs_list: List[Software], filename: str) -> BytesIO:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "new_sheet"
    init_file_columns(sheet)
    for row, program in enumerate(programs_list, 2):
        sheet.cell(row=row, column=1).value = program.name  # noqa WPS 110
        sheet.cell(row=row, column=2).value = program.vendor  # noqa WPS 110
        sheet.cell(row=row, column=3).value = program.version  # noqa WPS 110

    virtual_file = BytesIO(save_virtual_workbook(wb))
    virtual_file.name = filename
    return virtual_file
